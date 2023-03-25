import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import logging.config
import yaml
import numpy as np


with open('logging_conf.yaml', 'rt') as f:
    config = yaml.safe_load(f.read())
logging.config.dictConfig(config)
logger = logging.getLogger('data_transfer')

class DataTransfer:
    def __init__(self, input_file, target_file, date):
        self.input_file = input_file
        self.target_file = target_file
        self.date = date

    def transfer_data(self, input_names_row_range, input_values_row_range, target_names_row_range, input_values_col=1, input_names_col=0):

        input_df = pd.read_excel(self.input_file, sheet_name=None, engine='openpyxl')
        target_wb = load_workbook(self.target_file)

        for input_sheet_name, input_ws in input_df.items():
            input_sheet_name_stripped = input_sheet_name.strip()

            for target_ws in target_wb.worksheets:
                target_ws_title_stripped = target_ws.title.strip()

                if input_sheet_name_stripped == target_ws_title_stripped:
                    empty_col = target_ws.max_column

                    if target_ws.cell(row=6, column=empty_col).value != datetime.strptime(self.date, '%Y/%m/%d'):
                        empty_col += 1
                        target_ws.cell(row=6, column=empty_col, value=datetime.strptime(self.date, '%Y/%m/%d').date())

                    
                    input_names = input_ws.iloc[input_names_row_range[0]:input_names_row_range[1] + 1, input_names_col].values
                    # Replace Values that don't match input and target
                    input_names = np.where(input_names == "Total Privates", "Privates", input_names)
                    input_names = np.where(input_names == "Large Cap $5b+", "Large Cap", input_names)
                    input_names = np.where(input_names == "Mid Cap $1-5b", "Mid Cap", input_names)
                    input_names = np.where(input_names == "Small Cap <$1b", "Small Cap", input_names)
                    input_values =  input_ws.iloc[input_values_row_range[0]:input_values_row_range[1] + 1, input_values_col].values
                    
                    logger.debug(f"Processing sheet: {input_sheet_name}")
                    logger.debug("Input names: %s", input_names)
                    logger.debug("Input values: %s", input_values)
                    # print(f"Processing sheet: {input_sheet_name}")
                    # print("Input names:", input_names)
                    # print("Input values:", input_values)

                     # Filter out NaN values
                    input_data = [(name, value) for name, value in zip(input_names, input_values) if not (pd.isna(name) or pd.isna(value))]
                    # Testing
                    #-----------------------------------------------
                    # print(f"Processing sheet: {input_sheet_name}")
                    # print("Input data:", input_data)
                    # ----------------------------------------------
                    for i in range(target_names_row_range[0], target_names_row_range[1] + 1):
                      for col in range(2,5):
                        target_name = target_ws.cell(row=i, column=col).value
                        for input_name, input_value in input_data:
                            if input_name == target_name:
                                if not pd.isna(input_value):
                                  input_value = float(input_value)
                                  input_value_pct = f'{input_value * 100:.2f}%'
                                  target_ws.cell(row=i, column=empty_col, value=input_value_pct)
                                #   print(f"Mapping {input_name} ({input_value_pct}) to row {i}, column {empty_col}")
                                  logger.debug(f"Mapping {input_name} ({input_value_pct}) to row {i}, column {empty_col}")
                                  break

        target_wb.save(self.target_file)
   